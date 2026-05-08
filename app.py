import streamlit as st
import pandas as pd
import numpy as np
import math
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="발주량 산정 시스템",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');

html, body, [class*="css"] { font-family: 'Noto Sans KR', sans-serif; }

.main { background: #F8F9FB; }
[data-testid="stSidebar"] { background: #1A2236; }
[data-testid="stSidebar"] * { color: #C8D0E0 !important; }
[data-testid="stSidebar"] .sidebar-title { color: #FFFFFF !important; font-size: 18px; font-weight: 700; padding: 8px 0 4px; }
[data-testid="stSidebar"] hr { border-color: #2D3A52; }

.block-container { padding: 1.5rem 2rem; }

.kpi-card { background: white; border-radius: 12px; padding: 20px 24px; border: 1px solid #E8EDF5; }
.kpi-label { font-size: 12px; color: #8892A4; font-weight: 500; letter-spacing: .04em; margin-bottom: 6px; }
.kpi-value { font-size: 28px; font-weight: 700; line-height: 1; }
.kpi-sub { font-size: 12px; color: #8892A4; margin-top: 6px; }
.kpi-blue  { color: #2563EB; }
.kpi-amber { color: #D97706; }
.kpi-red   { color: #DC2626; }
.kpi-green { color: #059669; }

.section-header { font-size: 15px; font-weight: 600; color: #1E2A3B; margin-bottom: 12px; display: flex; align-items: center; gap: 8px; }

.badge-urgent  { background:#FEE2E2; color:#991B1B; padding:2px 8px; border-radius:99px; font-size:11px; font-weight:600; }
.badge-need    { background:#FEF3C7; color:#92400E; padding:2px 8px; border-radius:99px; font-size:11px; font-weight:600; }
.badge-ok      { background:#D1FAE5; color:#065F46; padding:2px 8px; border-radius:99px; font-size:11px; font-weight:600; }

.step-bar { display: flex; align-items: center; gap: 0; background: white; border-radius: 10px; padding: 14px 20px; border: 1px solid #E8EDF5; margin-bottom: 20px; }
.step-item { display: flex; align-items: center; gap: 8px; }
.step-num-done  { width:26px;height:26px;border-radius:50%;background:#059669;color:white;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700; }
.step-num-active{ width:26px;height:26px;border-radius:50%;background:#2563EB;color:white;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700; }
.step-num-idle  { width:26px;height:26px;border-radius:50%;background:#E8EDF5;color:#8892A4;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:600; }
.step-label-active { font-size:13px;font-weight:600;color:#2563EB; }
.step-label-done   { font-size:13px;font-weight:500;color:#059669; }
.step-label-idle   { font-size:13px;color:#8892A4; }
.step-line { flex:1; height:1px; background:#E8EDF5; margin:0 16px; }

.upload-box { background:white; border:2px dashed #CBD5E1; border-radius:12px; padding:28px; text-align:center; }
.upload-box-title { font-size:14px; font-weight:600; color:#1E2A3B; margin:10px 0 4px; }
.upload-box-sub { font-size:12px; color:#8892A4; }
</style>
""", unsafe_allow_html=True)


def n(v):
    try:
        return float(v) if pd.notna(v) and v != "" else 0.0
    except:
        return 0.0


def calc_order(row, month, default_safety_days=14, order_cycle=30):
    cpp = n(row.get("CPP"))
    if cpp <= 0:
        cpp = 1
    fc_map = {
        "2026-05": "예측수량_2026-05",
        "2026-06": "예측수량_2026-06",
        "2026-07": "예측수량_2026-07",
    }
    tg_map = {
        "2026-05": "기말재고_2026-05",
        "2026-06": "기말재고_2026-06",
        "2026-07": "기말재고_2026-07",
    }
    pk_map = {
        "2026-05": "기말예상재고_2026-04",
        "2026-06": "기말재고_2026-05",
        "2026-07": "기말재고_2026-06",
    }
    fc = n(row.get(fc_map.get(month, ""), 0))
    tg = n(row.get(tg_map.get(month, ""), 0))
    pv = n(row.get(pk_map.get(month, ""), 0))
    raw = fc + tg - pv
    if raw <= 0:
        return 0
    return math.ceil(raw / cpp) * int(cpp)


def build_result(df1, df2):
    df2r = df2.rename(
        columns={"단품코드": "품목코드", "단품컬러": "색상코드", "적재단위": "CPP", "브랜드구분": "브랜드"}
    )
    df = df1.merge(df2r[["품목코드", "색상코드", "CPP", "브랜드"]], on=["품목코드", "색상코드"], how="left")

    MONTHS = ["2026-05", "2026-06", "2026-07"]
    rows = []
    for _, row in df.iterrows():
        r = row.to_dict()
        for month in MONTHS:
            oqty = calc_order(r, month)
            pk_col = {
                "2026-05": "기말예상재고_2026-04",
                "2026-06": "기말재고_2026-05",
                "2026-07": "기말재고_2026-06",
            }
            prev = int(n(r.get(pk_col[month])))
            safety = int(n(r.get("안전재고", 0)))
            status = (
                "긴급"
                if oqty > 0 and prev <= safety
                else ("발주필요" if oqty > 0 else "여유")
            )
            cpp_v = r.get("CPP")
            cpp_out = int(n(cpp_v)) if pd.notna(cpp_v) and n(cpp_v) > 0 else None
            rows.append(
                {
                    "품목코드": str(r["품목코드"]) if pd.notna(r["품목코드"]) else "",
                    "색상코드": str(r["색상코드"]) if pd.notna(r["색상코드"]) else "",
                    "품목명": str(r.get("품목명", "")) if pd.notna(r.get("품목명")) else "",
                    "브랜드": str(r.get("브랜드", "")) if pd.notna(r.get("브랜드")) else "",
                    "공급처": str(r.get("공급처", "")) if pd.notna(r.get("공급처")) else "",
                    "사용구분": str(r.get("사용구분", "")) if pd.notna(r.get("사용구분")) else "",
                    "용도구분": str(r.get("용도구분", "")) if pd.notna(r.get("용도구분")) else "",
                    "12개월평균출고": round(n(r.get("12개월평균")), 1),
                    "발주월": month,
                    "출고예상": int(n(r.get(f"예측수량_{month}"))),
                    "직전월말예상재고": prev,
                    "월말목표재고": int(n(r.get(f"기말재고_{month}"))),
                    "안전재고": safety,
                    "적정재고": int(n(r.get("적정재고"))),
                    "CPP": cpp_out,
                    "권고발주량": oqty,
                    "발주후월말예상재고": prev + oqty - int(n(r.get(f"예측수량_{month}"))),
                    "소진가능개월수": round((prev + oqty - int(n(r.get(f"예측수량_{month}")))) / n(r.get("12개월평균")) , 1) if n(r.get("12개월평균")) > 0 else None,
                    "상태": status,
                }
            )
    return pd.DataFrame(rows)


def to_excel_bytes(rdf):
    filtered = rdf[rdf["권고발주량"] > 0].copy()
    HDR_FILL = PatternFill("solid", start_color="1F3864")
    URGENT = PatternFill("solid", start_color="FCE4D6")
    NEED = PatternFill("solid", start_color="FFF2CC")
    ALT = PatternFill("solid", start_color="F5F5F5")
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY = Font(name="Arial", size=9)
    BLUE_B = Font(name="Arial", bold=True, size=9, color="1F3864")
    thin = Side(style="thin", color="D0D0D0")
    BRD = Border(left=thin, right=thin, top=thin, bottom=thin)
    C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L = Alignment(horizontal="left", vertical="center")
    R = Alignment(horizontal="right", vertical="center")
    STATUS_F = {"긴급": URGENT, "발주필요": NEED}

    def h(ws, r, c, val):
        cl = ws.cell(row=r, column=c, value=val)
        cl.fill = HDR_FILL; cl.font = HDR_FONT; cl.alignment = C; cl.border = BRD

    def sc(ws, r, c, val, al=C, font=BODY, fill=None, fmt=None):
        cl = ws.cell(row=r, column=c, value=val)
        cl.alignment = al; cl.font = font; cl.border = BRD
        if fill: cl.fill = fill
        if fmt: cl.number_format = fmt
        return cl

    wb = Workbook()
    MONTHS = ["2026-05", "2026-06", "2026-07"]
    COLS = ["품목코드","색상코드","품목명","브랜드","공급처","사용구분","용도구분",
            "12개월평균출고","발주월","출고예상","직전월말예상재고","월말목표재고",
            "안전재고","적정재고","CPP","권고발주량","발주후월말예상재고","소진가능개월수","상태"]
    WIDS = [14,8,24,8,14,8,12,10,9,9,12,12,8,8,6,10,12,12,8]
    ALNS = [L,C,L,C,L,C,C,R,C,R,R,R,R,R,R,R,R,R,C]

    ws1 = wb.active; ws1.title = "발주량_통합"; ws1.freeze_panes = "J2"
    ws1.row_dimensions[1].height = 28
    for ci, (col, w) in enumerate(zip(COLS, WIDS), 1):
        h(ws1, 1, ci, col)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    for ri, (_, row) in enumerate(filtered.iterrows(), 2):
        bg = ALT if ri % 2 == 0 else None
        st_val = row["상태"]
        for ci, (col, al) in enumerate(zip(COLS, ALNS), 1):
            val = row[col]
            if val is None or (isinstance(val, float) and math.isnan(val)):
                val = ""
            elif isinstance(val, float) and col not in ["12개월평균출고"] and val == int(val):
                val = int(val)
            is_num = isinstance(val, (int, float)) and col not in ["CPP"]
            fmt = "#,##0.0" if col in ["12개월평균출고","소진가능개월수"] else ("#,##0" if is_num and val != "" else None)
            fill = STATUS_F.get(st_val) if col == "상태" else bg
            fnt = BLUE_B if col == "권고발주량" else BODY
            sc(ws1, ri, ci, val, al=al, font=fnt, fill=fill, fmt=fmt)
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    ws2 = wb.create_sheet("월별_요약"); ws2.row_dimensions[1].height = 28
    for ci, (c, w) in enumerate(zip(["발주월","전체품목수","발주필요","긴급","총권고발주량","공급처수"],[10,12,12,10,14,10]), 1):
        h(ws2, 1, ci, c); ws2.column_dimensions[get_column_letter(ci)].width = w
    for ri, month in enumerate(MONTHS, 2):
        m_all = rdf[rdf["발주월"] == month]
        m_ord = m_all[m_all["권고발주량"] > 0]
        for ci, v in enumerate([month, len(m_all), len(m_ord), int((m_ord["상태"]=="긴급").sum()),
                                  int(m_ord["권고발주량"].sum()), int(m_ord["공급처"].nunique())], 1):
            sc(ws2, ri, ci, v, fmt="#,##0" if isinstance(v, int) and ci > 1 else None)

    ws3 = wb.create_sheet("공급처별_발주"); ws3.row_dimensions[1].height = 28
    for ci, (c, w) in enumerate(zip(["공급처","발주월","발주품목수","긴급건수","총권고발주량"],[18,10,10,10,14]), 1):
        h(ws3, 1, ci, c); ws3.column_dimensions[get_column_letter(ci)].width = w
    ri = 2
    for month in MONTHS:
        m_ord = filtered[filtered["발주월"] == month]
        for sup, g in m_ord.groupby("공급처", sort=True):
            bg = ALT if ri % 2 == 0 else None
            for ci, v in enumerate([sup, month, len(g), int((g["상태"]=="긴급").sum()), int(g["권고발주량"].sum())], 1):
                sc(ws3, ri, ci, v, al=L if ci==1 else C, fill=bg, fmt="#,##0" if isinstance(v, int) and ci > 2 else None)
            ri += 1
    ws3.auto_filter.ref = "A1:E1"

    ws4 = wb.create_sheet("발주서_공급처별"); ws4.row_dimensions[1].height = 28
    ORDER_COLS = ["공급처","품목코드","색상코드","품목명","브랜드","발주월","출고예상","권고발주량","발주후월말예상재고","소진가능개월수","CPP"]
    ORDER_W    = [16,14,8,24,8,9,10,10,12,10,7]
    for ci, (c, w) in enumerate(zip(ORDER_COLS, ORDER_W), 1):
        h(ws4, 1, ci, c); ws4.column_dimensions[get_column_letter(ci)].width = w
    order_data = filtered.sort_values(["공급처","발주월"])
    for ri, (_, row) in enumerate(order_data.iterrows(), 2):
        bg = ALT if ri % 2 == 0 else None
        for ci, col in enumerate(ORDER_COLS, 1):
            val = row[col]
            if val is None or (isinstance(val, float) and math.isnan(val)): val = ""
            is_num = isinstance(val, (int, float)) and col not in ["CPP"]
            fmt = "#,##0" if is_num and val != "" else None
            sc(ws4, ri, ci, val, al=L if col in ["공급처","품목명"] else C, fill=bg, fmt=fmt)
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(ORDER_COLS))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── 사이드바 ──────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="sidebar-title">📦 발주량 산정 시스템</div>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("**메뉴**")
    menu = st.radio(
        label="",
        options=["🏠 홈 / 파일 업로드", "📊 발주량 결과", "📋 발주서 보기", "📈 공급처별 요약"],
        label_visibility="collapsed",
    )
    st.markdown("---")
    st.markdown("**파라미터 설정**")
    default_safety = st.number_input("기본 안전재고일수", min_value=0, max_value=90, value=14, step=1)
    order_cycle = st.number_input("발주 주기 (일)", min_value=1, max_value=90, value=30, step=1)
    st.markdown("---")
    st.caption("© 2026 발주관리팀")


# ── 세션 상태 ──────────────────────────────────────────────
if "result_df" not in st.session_state:
    st.session_state.result_df = None


# ── 홈 / 파일 업로드 ──────────────────────────────────────
if menu == "🏠 홈 / 파일 업로드":
    st.markdown("## 파일 업로드")
    st.markdown("두 파일을 업로드하면 자동으로 발주량을 계산합니다.")

    # 진행 단계 표시
    has_result = st.session_state.result_df is not None
    step_html = f"""
    <div class="step-bar">
      <div class="step-item">
        <div class="step-num-{'done' if has_result else 'active'}">{'✓' if has_result else '1'}</div>
        <span class="step-label-{'done' if has_result else 'active'}">파일 업로드</span>
      </div>
      <div class="step-line"></div>
      <div class="step-item">
        <div class="step-num-{'active' if has_result else 'idle'}">2</div>
        <span class="step-label-{'active' if has_result else 'idle'}">발주량 계산</span>
      </div>
      <div class="step-line"></div>
      <div class="step-item">
        <div class="step-num-idle">3</div>
        <span class="step-label-idle">발주서 변환</span>
      </div>
      <div class="step-line"></div>
      <div class="step-item">
        <div class="step-num-idle">4</div>
        <span class="step-label-idle">다운로드</span>
      </div>
    </div>
    """
    st.markdown(step_html, unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("#### 📄 수요예측 파일")
        file1 = st.file_uploader(
            "수요예측 xlsx 업로드",
            type=["xlsx", "xls"],
            key="file1",
            label_visibility="collapsed",
        )
        if file1:
            st.success(f"✅ {file1.name} 업로드 완료")

    with col2:
        st.markdown("#### 📄 CPP(적재단위) 파일")
        file2 = st.file_uploader(
            "CPP xls 업로드",
            type=["xlsx", "xls"],
            key="file2",
            label_visibility="collapsed",
        )
        if file2:
            st.success(f"✅ {file2.name} 업로드 완료")

    st.markdown("---")

    if file1 and file2:
        if st.button("🚀 발주량 계산 시작", type="primary", use_container_width=True):
            with st.spinner("데이터를 읽고 발주량을 계산하는 중..."):
                try:
                    from openpyxl import load_workbook
                    wb_src = load_workbook(file1, read_only=True)
                    ws_src = wb_src["품목별상세"]
                    rows = list(ws_src.iter_rows(values_only=True))
                    df1 = pd.DataFrame(rows[1:], columns=rows[0])

                    df2 = pd.read_excel(file2, engine="xlrd" if str(file2.name).endswith(".xls") else "openpyxl")
                    df2 = df2.rename(columns={"단품코드": "품목코드", "단품컬러": "색상코드", "적재단위": "CPP", "브랜드구분": "브랜드"})

                    rdf = build_result(df1, df2)
                    st.session_state.result_df = rdf
                    st.success(f"✅ 계산 완료! 총 {len(rdf[rdf['권고발주량']>0]):,}건의 발주가 필요합니다.")
                    st.balloons()
                except Exception as e:
                    st.error(f"오류가 발생했습니다: {e}")
    else:
        st.info("두 파일을 모두 업로드하면 계산 버튼이 활성화됩니다.")

    if st.session_state.result_df is not None:
        rdf = st.session_state.result_df
        st.markdown("---")
        st.markdown("### 📥 결과 다운로드")
        col_a, col_b = st.columns(2)
        with col_a:
            excel_bytes = to_excel_bytes(rdf)
            st.download_button(
                label="📊 발주량 결과 Excel 다운로드",
                data=excel_bytes,
                file_name=f"발주량_산정결과_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_b:
            st.download_button(
                label="📋 발주서 CSV 다운로드",
                data=rdf[rdf["권고발주량"]>0].to_csv(index=False, encoding="utf-8-sig"),
                file_name=f"발주서_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True,
            )


# ── 발주량 결과 ──────────────────────────────────────────────
elif menu == "📊 발주량 결과":
    st.markdown("## 발주량 결과")

    if st.session_state.result_df is None:
        st.warning("먼저 홈에서 파일을 업로드하고 계산을 실행해 주세요.")
        st.stop()

    rdf = st.session_state.result_df
    filtered = rdf[rdf["권고발주량"] > 0]
    MONTHS = ["2026-05", "2026-06", "2026-07"]

    # KPI
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"""<div class="kpi-card"><div class="kpi-label">전체 품목 (건수)</div>
        <div class="kpi-value kpi-blue">{len(rdf):,}</div>
        <div class="kpi-sub">3개월 합산</div></div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"""<div class="kpi-card"><div class="kpi-label">발주 필요</div>
        <div class="kpi-value kpi-amber">{len(filtered):,}</div>
        <div class="kpi-sub">전체의 {len(filtered)/len(rdf)*100:.1f}%</div></div>""", unsafe_allow_html=True)
    with col3:
        urgent_cnt = int((filtered["상태"]=="긴급").sum())
        st.markdown(f"""<div class="kpi-card"><div class="kpi-label">긴급 발주</div>
        <div class="kpi-value kpi-red">{urgent_cnt:,}</div>
        <div class="kpi-sub">재고 ≤ 안전재고</div></div>""", unsafe_allow_html=True)
    with col4:
        sup_cnt = filtered["공급처"].nunique()
        st.markdown(f"""<div class="kpi-card"><div class="kpi-label">발주 공급처</div>
        <div class="kpi-value kpi-green">{sup_cnt}</div>
        <div class="kpi-sub">개 공급처</div></div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # 월별 요약 테이블
    st.markdown("#### 월별 발주 요약")
    summary_rows = []
    for month in MONTHS:
        m = rdf[rdf["발주월"]==month]
        m_o = m[m["권고발주량"]>0]
        summary_rows.append({
            "발주월": month,
            "전체 품목수": f"{len(m):,}",
            "발주필요": f"{len(m_o):,}",
            "긴급": f"{int((m_o['상태']=='긴급').sum()):,}",
            "총 권고발주량": f"{int(m_o['권고발주량'].sum()):,}",
            "공급처 수": f"{m_o['공급처'].nunique()}",
        })
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

    st.markdown("---")

    # 상세 필터 테이블
    st.markdown("#### 발주 상세 내역")
    c1, c2, c3 = st.columns([2, 2, 2])
    with c1:
        month_filter = st.selectbox("발주월", ["전체"] + MONTHS)
    with c2:
        status_filter = st.selectbox("상태", ["전체", "긴급", "발주필요", "여유"])
    with c3:
        brand_filter = st.selectbox("브랜드", ["전체"] + sorted(rdf["브랜드"].dropna().unique().tolist()))

    disp = rdf.copy()
    if month_filter != "전체":
        disp = disp[disp["발주월"]==month_filter]
    if status_filter != "전체":
        disp = disp[disp["상태"]==status_filter]
    if brand_filter != "전체":
        disp = disp[disp["브랜드"]==brand_filter]

    show_cols = ["품목코드","색상코드","품목명","브랜드","공급처","발주월","출고예상","직전월말예상재고","월말목표재고","안전재고","CPP","권고발주량","발주후월말예상재고","소진가능개월수","상태"]
    st.dataframe(
        disp[show_cols].reset_index(drop=True),
        use_container_width=True,
        hide_index=True,
        column_config={
            "권고발주량": st.column_config.NumberColumn(format="%d"),
            "출고예상": st.column_config.NumberColumn(format="%d"),
            "발주후월말예상재고": st.column_config.NumberColumn(format="%d"),
            "소진가능개월수": st.column_config.NumberColumn(format="%.1f"),
        }
    )
    st.caption(f"총 {len(disp):,}건 표시 중")


# ── 발주서 보기 ──────────────────────────────────────────────
elif menu == "📋 발주서 보기":
    st.markdown("## 발주서 (공급처별)")

    if st.session_state.result_df is None:
        st.warning("먼저 홈에서 파일을 업로드하고 계산을 실행해 주세요.")
        st.stop()

    rdf = st.session_state.result_df
    filtered = rdf[rdf["권고발주량"] > 0].copy()

    MONTHS = ["2026-05", "2026-06", "2026-07"]
    sel_month = st.selectbox("발주월 선택", MONTHS)
    month_data = filtered[filtered["발주월"]==sel_month]

    suppliers = sorted(month_data["공급처"].dropna().unique().tolist())
    sel_sup = st.selectbox("공급처 선택", ["전체"] + suppliers)

    if sel_sup != "전체":
        month_data = month_data[month_data["공급처"]==sel_sup]

    st.markdown("---")
    st.markdown(f"#### {sel_month} 발주서 — {sel_sup}")

    order_cols = ["품목코드","색상코드","품목명","브랜드","공급처","출고예상","직전월말예상재고","월말목표재고","CPP","권고발주량"]
    st.dataframe(
        month_data[order_cols].sort_values("공급처").reset_index(drop=True),
        use_container_width=True,
        hide_index=True,
        column_config={"권고발주량": st.column_config.NumberColumn("권고발주량", format="%d")}
    )
    st.caption(f"총 {len(month_data):,}건 / 권고발주량 합계: {int(month_data['권고발주량'].sum()):,}")

    st.markdown("---")
    excel_bytes = to_excel_bytes(rdf)
    st.download_button(
        label=f"📥 {sel_month} 발주서 Excel 다운로드",
        data=excel_bytes,
        file_name=f"발주서_{sel_month}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ── 공급처별 요약 ──────────────────────────────────────────────
elif menu == "📈 공급처별 요약":
    st.markdown("## 공급처별 발주 요약")

    if st.session_state.result_df is None:
        st.warning("먼저 홈에서 파일을 업로드하고 계산을 실행해 주세요.")
        st.stop()

    rdf = st.session_state.result_df
    filtered = rdf[rdf["권고발주량"] > 0].copy()

    MONTHS = ["2026-05", "2026-06", "2026-07"]
    sel_month = st.selectbox("발주월", ["전체"] + MONTHS)
    if sel_month != "전체":
        filtered = filtered[filtered["발주월"]==sel_month]

    sup_summary = (
        filtered.groupby("공급처")
        .agg(
            발주품목수=("품목코드", "count"),
            긴급건수=("상태", lambda x: (x=="긴급").sum()),
            총권고발주량=("권고발주량", "sum"),
        )
        .reset_index()
        .sort_values("총권고발주량", ascending=False)
    )
    sup_summary["긴급비율"] = (sup_summary["긴급건수"] / sup_summary["발주품목수"] * 100).round(1).astype(str) + "%"

    st.dataframe(sup_summary, use_container_width=True, hide_index=True,
        column_config={
            "총권고발주량": st.column_config.NumberColumn(format="%d"),
            "발주품목수": st.column_config.NumberColumn(format="%d"),
            "긴급건수": st.column_config.NumberColumn(format="%d"),
        }
    )
    st.caption(f"공급처 {len(sup_summary)}개")

    st.markdown("---")
    st.markdown("#### 공급처별 발주량 상위 20개")
    chart_data = sup_summary.head(20).set_index("공급처")["총권고발주량"]
    st.bar_chart(chart_data)
